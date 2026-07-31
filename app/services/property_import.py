"""Bulk import of an agency's property catalogue. TZ 14 (objects feed matching).

Production ran on two seed objects, so matching, pitches, commercial offers and
the funnel analytics all had nothing real to work with -- the buyer half of the
system was finished while the inventory half was empty.

Agencies keep their catalogue in Excel or a CRM export, with Russian headers and
human-formatted values ("8 500 000 ₽", "56,3", "2-комн", "да"). This module maps
those onto the Property model, coerces the values, and reports per-row problems
instead of failing the whole file.

Import is idempotent per source_url when present, otherwise per
(title, address) within the agency, so re-uploading a corrected file updates
rows rather than duplicating the catalogue.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any, Iterable, Optional

import structlog

logger = structlog.get_logger()

# migrations/001_init.sql CHECK constraints
PROPERTY_TYPES = {"apartment", "house", "commercial", "land", "studio"}
READINESS = {"ready", "under_construction", "foundation", "planning"}
STATUSES = {"active", "sold", "reserved", "archive"}

# Header synonyms -> model field. Compared case-insensitively with punctuation
# and spaces stripped, so "Цена, ₽" and "цена руб" both land on `price`.
COLUMN_MAP: dict[str, str] = {
    "title": "title", "название": "title", "наименование": "title", "объект": "title",
    "propertytype": "property_type", "тип": "property_type", "типобъекта": "property_type",
    "price": "price", "цена": "price", "стоимость": "price",
    "areatotal": "area_total", "площадь": "area_total", "общаяплощадь": "area_total",
    "arealiving": "area_living", "жилаяплощадь": "area_living",
    "rooms": "rooms", "комнат": "rooms", "количествокомнат": "rooms", "комнаты": "rooms",
    "floor": "floor", "этаж": "floor",
    "floorstotal": "floors_total", "этажей": "floors_total", "этажность": "floors_total",
    "district": "district", "район": "district",
    "address": "address", "адрес": "address",
    "developer": "developer", "застройщик": "developer",
    "yearbuilt": "year_built", "годпостройки": "year_built", "год": "year_built",
    "isnewbuild": "is_new_build", "новостройка": "is_new_build",
    "readinessstatus": "readiness_status", "готовность": "readiness_status",
    "status": "status", "статус": "status",
    "description": "description_original", "описание": "description_original",
    "sourceurl": "source_url", "ссылка": "source_url", "url": "source_url",
    "pricepersqm": "price_per_sqm", "ценазам2": "price_per_sqm",
}

_TYPE_WORDS = {
    "apartment": ("квартир", "апартамент", "apartment", "flat"),
    "studio": ("студи", "studio"),
    "house": ("дом", "коттедж", "таунхаус", "house", "cottage"),
    "land": ("участок", "земл", "land", "plot"),
    "commercial": ("коммерч", "офис", "магазин", "помещен", "commercial"),
}
_READINESS_WORDS = {
    "ready": ("готов", "сдан", "ready", "введ"),
    "under_construction": ("строит", "construction", "стройк"),
    "foundation": ("котлован", "фундамент", "foundation"),
    "planning": ("проект", "planning", "план"),
}
_STATUS_WORDS = {
    "sold": ("продан", "sold"),
    "reserved": ("бронь", "резерв", "reserved"),
    "archive": ("архив", "archive", "снят"),
    "active": ("актив", "active", "в продаже", "продажа"),
}
_TRUE = ("да", "yes", "true", "1", "+", "новостройка")
_FALSE = ("нет", "no", "false", "0", "-", "вторичка", "вторичное")

_MULTIPLIERS = (("млрд", 1_000_000_000), ("млн", 1_000_000), ("тыс", 1_000))


def normalize_header(name: str) -> str:
    return re.sub(r"[^a-zа-я0-9]", "", str(name or "").strip().lower())


def parse_number(value: Any) -> Optional[float]:
    """Read a human-written number: "8 500 000 ₽", "8,5 млн", "56,3", "1 200"."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().lower().replace("\xa0", " ")
    if not text:
        return None

    multiplier = 1
    for word, factor in _MULTIPLIERS:
        if word in text:
            multiplier = factor
            text = text.replace(word, " ")
            break

    # Keep digits, separators and the sign; drop currency and stray words.
    text = re.sub(r"[^\d,.\-]", "", text)
    # A "-" only means a sign at the very front. Anywhere else it is a leftover
    # from "2-комн" or a range, and float() would choke on it.
    sign = -1 if text.startswith("-") else 1
    text = text.lstrip("-").replace("-", "")
    # Trailing punctuation survives from things like "руб." and would otherwise
    # be read as a thousands separator, inflating the value a hundredfold.
    text = text.strip(".,")
    if not text:
        return None
    # A comma is a decimal separator in Russian exports; a dot may be either, and
    # thousands are usually spaces (already gone).
    text = text.replace(",", ".")
    if text.count(".") > 1:  # "8.500.000" -> thousands separators
        text = text.replace(".", "")
    try:
        return float(text) * multiplier * sign
    except ValueError:
        return None


def parse_int(value: Any) -> Optional[int]:
    num = parse_number(value)
    return int(round(num)) if num is not None else None


def parse_bool(value: Any) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if not text:
        return None
    if any(t == text or text.startswith(t) for t in _TRUE):
        return True
    if any(t == text or text.startswith(t) for t in _FALSE):
        return False
    return None


def parse_rooms(value: Any) -> Optional[int]:
    """"2", "2-комн", "двухкомнатная", "студия" -> 2, 2, 2, 0."""
    text = str(value or "").strip().lower()
    if not text:
        return None
    if "студи" in text:
        return 0
    words = {"одно": 1, "двух": 2, "трех": 3, "трёх": 3, "четырех": 4, "четырёх": 4, "пяти": 5}
    for word, n in words.items():
        if word in text:
            return n
    return parse_int(text)


def _match_word(value: Any, table: dict[str, tuple], default: Optional[str] = None) -> Optional[str]:
    text = str(value or "").strip().lower()
    if not text:
        return default
    for key, words in table.items():
        if key == text or any(w in text for w in words):
            return key
    return default


def parse_property_type(value: Any) -> Optional[str]:
    return _match_word(value, _TYPE_WORDS)


def parse_readiness(value: Any) -> Optional[str]:
    return _match_word(value, _READINESS_WORDS)


def parse_status(value: Any) -> str:
    return _match_word(value, _STATUS_WORDS, default="active") or "active"


@dataclass
class RowError:
    row: int
    message: str


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[RowError] = field(default_factory=list)
    unmapped_columns: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": [{"row": e.row, "message": e.message} for e in self.errors],
            "unmapped_columns": self.unmapped_columns,
        }


def read_rows(content: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """Return (rows, headers) from a CSV or XLSX payload."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _read_xlsx(content)
    return _read_csv(content)


def _read_csv(content: bytes) -> tuple[list[dict], list[str]]:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Не удалось определить кодировку файла (ожидается UTF-8 или CP1251)")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return list(reader), list(reader.fieldnames or [])


def _read_xlsx(content: bytes) -> tuple[list[dict], list[str]]:
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError as e:
        raise ValueError("Чтение XLSX требует зависимости openpyxl") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return [], []
    out = [
        dict(zip(headers, row))
        for row in rows
        if any(cell is not None and str(cell).strip() for cell in row)
    ]
    wb.close()
    return out, headers


def map_row(row: dict, headers: list[str]) -> dict:
    """Map one spreadsheet row onto Property fields (unvalidated)."""
    values: dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        field_name = COLUMN_MAP.get(normalize_header(raw_key))
        if field_name and (raw_value is not None and str(raw_value).strip() != ""):
            values[field_name] = raw_value

    out: dict[str, Any] = {}
    if "title" in values:
        out["title"] = str(values["title"]).strip()
    for key in ("address", "district", "developer", "source_url", "description_original"):
        if key in values:
            out[key] = str(values[key]).strip()
    for key in ("price", "price_per_sqm", "floor", "floors_total", "year_built"):
        if key in values:
            out[key] = parse_int(values[key])
    for key in ("area_total", "area_living"):
        if key in values:
            out[key] = parse_number(values[key])
    if "rooms" in values:
        out["rooms"] = parse_rooms(values["rooms"])
    if "property_type" in values:
        out["property_type"] = parse_property_type(values["property_type"])
    if "readiness_status" in values:
        out["readiness_status"] = parse_readiness(values["readiness_status"])
    if "is_new_build" in values:
        out["is_new_build"] = parse_bool(values["is_new_build"])
    out["status"] = parse_status(values.get("status"))

    # A studio has no separate room count; keep the type consistent with it.
    if out.get("rooms") == 0 and not out.get("property_type"):
        out["property_type"] = "studio"
    # Derive the per-sqm price when both sides are known; matching and the
    # market analytics both read it.
    if not out.get("price_per_sqm") and out.get("price") and out.get("area_total"):
        out["price_per_sqm"] = int(out["price"] / out["area_total"])

    return {k: v for k, v in out.items() if v is not None}


def validate_row(mapped: dict) -> Optional[str]:
    """Return an error message, or None when the row is importable."""
    if not mapped.get("title"):
        return "нет названия объекта"
    price = mapped.get("price")
    if price is None:
        return "нет цены"
    if price <= 0:
        return "цена должна быть больше нуля"
    if price < 100_000:
        # Catches a catalogue written in thousands, which would silently make
        # every object match every budget.
        return f"цена подозрительно мала ({price}) — возможно, указана в тысячах"
    if mapped.get("property_type") and mapped["property_type"] not in PROPERTY_TYPES:
        return f"неизвестный тип объекта: {mapped['property_type']}"
    if mapped.get("readiness_status") and mapped["readiness_status"] not in READINESS:
        return f"неизвестная готовность: {mapped['readiness_status']}"
    if mapped.get("status") not in STATUSES:
        return f"неизвестный статус: {mapped.get('status')}"
    return None


def unmapped_columns(headers: Iterable[str]) -> list[str]:
    return [h for h in headers if h and normalize_header(h) not in COLUMN_MAP]
